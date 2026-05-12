#!/usr/bin/env python3
"""Visualise an EPA SWMM .inp network in 3D.

The script reads junctions, outfalls, conduits and xsections from an INP file,
reads node XY coordinates from an Excel file, then builds a 3D view where:

* Junctions are black nodes.
* Outfalls are red nodes.
* Node coordinates are X = Excel column B, Y = Excel column C, Z = INP Elevation.
* Node shapes are read from Excel column D: Fixed = cube, NotFixed = sphere.
* Conduits are rendered from their SWMM cross-section geometry and colored by
  the ratio between real 3D node distance and INP Length.

Plotly is used for the interactive HTML output. If Plotly is not installed, the
script can still print a summary of the parsed network.

Spyder / Anaconda usage
----------------------
Edit the configuration block below, then run this file directly from Spyder.
"""

from __future__ import annotations

import argparse
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


SCRIPT_DIR = Path(__file__).resolve().parent


# =============================================================================
# Configuration Spyder
# =============================================================================
# Change only these values when running the script directly from Spyder.
INP_FILE = "SWMM_Twannbach.inp"
COORDINATES_XLSX = "260508_Coord_nodes_SWMM.xlsx"
OUTPUT_HTML = ""  # Empty = automatic name: <INP_FILE>_3d.html
OUTPUT_OBJ = ""  # Empty = automatic name: <INP_FILE>_network.obj
EXPORT_OBJ = True
OBJ_SWAP_YZ = True  # Write OBJ coordinates as X, Elevation, Y.
OBJ_EXPORT_NODES = True
OBJ_NODE_RADIUS = 1.5
DIAMETER_SCALE = 1.0  # Use 1.0 for real conduit dimensions.
CROSS_SECTION_SEGMENTS = 16
SUMMARY_ONLY = False


@dataclass(frozen=True)
class Node:
    name: str
    kind: str
    elevation: float
    x: float | None = None
    y: float | None = None
    node_type: str = "NotFixed"

    @property
    def has_xyz(self) -> bool:
        return self.x is not None and self.y is not None


@dataclass(frozen=True)
class Conduit:
    name: str
    from_node: str
    to_node: str
    length: float
    shape: str = "UNKNOWN"
    geom1: float = 0.0
    geom2: float = 0.0
    geom3: float = 0.0


def strip_comment(line: str) -> str:
    return line.split(";", 1)[0].strip()


def to_float(value: str, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def read_sections(inp_path: Path) -> dict[str, list[str]]:
    sections: dict[str, list[str]] = {}
    current: str | None = None

    with inp_path.open("r", encoding="utf-8-sig", errors="replace") as handle:
        for raw_line in handle:
            line = raw_line.rstrip("\n")
            stripped = line.strip()
            if not stripped:
                continue
            if stripped.startswith("[") and stripped.endswith("]"):
                current = stripped[1:-1].upper()
                sections.setdefault(current, [])
                continue
            if current is not None:
                sections[current].append(line)

    return sections


def data_lines(lines: Iterable[str]) -> Iterable[list[str]]:
    for line in lines:
        clean = strip_comment(line)
        if clean:
            yield clean.split()


def read_coordinates_xlsx(xlsx_path: Path) -> dict[str, tuple[float, float, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read node coordinates from .xlsx files.") from exc

    coordinates: dict[str, tuple[float, float, str]] = {}
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=1, values_only=True):
        if len(row) < 3 or row[0] is None:
            continue
        name = str(row[0]).strip()
        if not name:
            continue
        x = to_float(str(row[1]), default=float("nan"))
        y = to_float(str(row[2]), default=float("nan"))
        if math.isnan(x) or math.isnan(y):
            continue
        node_type = "NotFixed"
        if len(row) >= 4 and row[3] is not None:
            node_type = str(row[3]).strip() or "NotFixed"
        coordinates[name] = (x, y, node_type)

    workbook.close()
    return coordinates


def parse_inp(inp_path: Path, coordinates_xlsx_path: Path | None = None) -> tuple[dict[str, Node], list[Conduit]]:
    sections = read_sections(inp_path)
    nodes: dict[str, Node] = {}

    for row in data_lines(sections.get("JUNCTIONS", [])):
        if len(row) >= 2:
            nodes[row[0]] = Node(name=row[0], kind="junction", elevation=to_float(row[1]))

    for row in data_lines(sections.get("OUTFALLS", [])):
        if len(row) >= 2:
            nodes[row[0]] = Node(name=row[0], kind="outfall", elevation=to_float(row[1]))

    if coordinates_xlsx_path is None:
        coordinates: dict[str, tuple[float, float, str]] = {}
        for row in data_lines(sections.get("COORDINATES", [])):
            if len(row) >= 3:
                coordinates[row[0]] = (to_float(row[1]), to_float(row[2]), "NotFixed")
    else:
        coordinates = read_coordinates_xlsx(coordinates_xlsx_path)

    for name, node in list(nodes.items()):
        if name in coordinates:
            x, y, node_type = coordinates[name]
            nodes[name] = Node(
                name=node.name,
                kind=node.kind,
                elevation=node.elevation,
                x=x,
                y=y,
                node_type=node_type,
            )

    xsections: dict[str, tuple[str, float, float, float]] = {}
    for row in data_lines(sections.get("XSECTIONS", [])):
        if len(row) >= 5:
            xsections[row[0]] = (
                row[1].upper(),
                to_float(row[2]),
                to_float(row[3]),
                to_float(row[4]),
            )

    conduits: list[Conduit] = []
    for row in data_lines(sections.get("CONDUITS", [])):
        if len(row) >= 4:
            shape, geom1, geom2, geom3 = xsections.get(row[0], ("UNKNOWN", 0.0, 0.0, 0.0))
            conduits.append(
                Conduit(
                    name=row[0],
                    from_node=row[1],
                    to_node=row[2],
                    length=to_float(row[3]),
                    shape=shape,
                    geom1=geom1,
                    geom2=geom2,
                    geom3=geom3,
                )
            )

    return nodes, conduits


def vector_sub(a: tuple[float, float, float], b: tuple[float, float, float]) -> tuple[float, float, float]:
    return (a[0] - b[0], a[1] - b[1], a[2] - b[2])


def vector_cross(a: tuple[float, float, float], b: tuple[float, float, float]) -> tuple[float, float, float]:
    return (
        a[1] * b[2] - a[2] * b[1],
        a[2] * b[0] - a[0] * b[2],
        a[0] * b[1] - a[1] * b[0],
    )


def vector_norm(a: tuple[float, float, float]) -> float:
    return math.sqrt(a[0] * a[0] + a[1] * a[1] + a[2] * a[2])


def vector_scale(a: tuple[float, float, float], factor: float) -> tuple[float, float, float]:
    return (a[0] * factor, a[1] * factor, a[2] * factor)


def vector_add(a: tuple[float, float, float], b: tuple[float, float, float]) -> tuple[float, float, float]:
    return (a[0] + b[0], a[1] + b[1], a[2] + b[2])


def unit(a: tuple[float, float, float]) -> tuple[float, float, float]:
    norm = vector_norm(a)
    if norm == 0:
        return (0.0, 0.0, 0.0)
    return vector_scale(a, 1.0 / norm)


def conduit_cross_section(conduit: Conduit, scale: float, segments: int) -> list[tuple[float, float]]:
    shape = conduit.shape.upper()

    if shape == "CIRCULAR":
        radius = max(conduit.geom1 * scale / 2.0, 0.001)
        return [
            (radius * math.cos(2.0 * math.pi * i / segments), radius * math.sin(2.0 * math.pi * i / segments))
            for i in range(segments)
        ]

    if shape == "RECT_CLOSED":
        width = max(conduit.geom1 * scale, 0.001)
        height = max(conduit.geom2 * scale, 0.001)
        return [
            (-width / 2.0, -height / 2.0),
            (width / 2.0, -height / 2.0),
            (width / 2.0, height / 2.0),
            (-width / 2.0, height / 2.0),
        ]

    if shape == "MODBASKETHANDLE":
        width = max(conduit.geom1 * scale, 0.001)
        height = max((conduit.geom2 or conduit.geom1) * scale, 0.001)
        return [
            (width / 2.0 * math.cos(2.0 * math.pi * i / segments), height / 2.0 * math.sin(2.0 * math.pi * i / segments))
            for i in range(segments)
        ]

    diameter = max(conduit.geom1 * scale, 0.001)
    radius = diameter / 2.0
    return [
        (radius * math.cos(2.0 * math.pi * i / segments), radius * math.sin(2.0 * math.pi * i / segments))
        for i in range(segments)
    ]


def tube_mesh(
    start: tuple[float, float, float],
    end: tuple[float, float, float],
    cross_section: list[tuple[float, float]],
) -> tuple[list[float], list[float], list[float], list[int], list[int], list[int]]:
    axis = unit(vector_sub(end, start))
    if axis == (0.0, 0.0, 0.0):
        return [], [], [], [], [], []

    reference = (0.0, 0.0, 1.0)
    if abs(axis[2]) > 0.95:
        reference = (0.0, 1.0, 0.0)

    u = unit(vector_cross(axis, reference))
    v = unit(vector_cross(axis, u))

    x: list[float] = []
    y: list[float] = []
    z: list[float] = []
    for center in (start, end):
        for cx, cy in cross_section:
            point = vector_add(center, vector_add(vector_scale(u, cx), vector_scale(v, cy)))
            x.append(point[0])
            y.append(point[1])
            z.append(point[2])

    n = len(cross_section)
    i_idx: list[int] = []
    j_idx: list[int] = []
    k_idx: list[int] = []

    for idx in range(n):
        next_idx = (idx + 1) % n
        i_idx.extend([idx, idx])
        j_idx.extend([next_idx, next_idx + n])
        k_idx.extend([idx + n, idx + n])

    return x, y, z, i_idx, j_idx, k_idx


def node_xyz(node: Node) -> tuple[float, float, float]:
    if node.x is None or node.y is None:
        raise ValueError(f"Node {node.name!r} has no coordinates")
    return (node.x, node.y, node.elevation)


def conduit_distance_ratio(conduit: Conduit, nodes: dict[str, Node]) -> float | None:
    start_node = nodes.get(conduit.from_node)
    end_node = nodes.get(conduit.to_node)
    if start_node is None or end_node is None or not start_node.has_xyz or not end_node.has_xyz:
        return None
    if conduit.length <= 0:
        return None

    distance = vector_norm(vector_sub(node_xyz(end_node), node_xyz(start_node)))
    return distance / conduit.length


def rgb_to_hex(rgb: tuple[int, int, int]) -> str:
    return "#{:02X}{:02X}{:02X}".format(*rgb)


def conduit_color_rgb(ratio: float | None) -> tuple[int, int, int]:
    if ratio is None:
        return (130, 130, 130)
    if 0.6 <= ratio <= 1.4:
        return (145, 145, 145)
    if ratio < 0.6:
        intensity = min((0.6 - ratio) / 0.6, 1.0)
        green_blue = round(145 * (1.0 - intensity))
        return (255, green_blue, green_blue)

    intensity = min((ratio - 1.4) / 1.4, 1.0)
    red_green = round(145 * (1.0 - intensity))
    return (red_green, red_green, 255)


def conduit_color_hex(conduit: Conduit, nodes: dict[str, Node]) -> str:
    return rgb_to_hex(conduit_color_rgb(conduit_distance_ratio(conduit, nodes)))


def build_plotly_figure(
    nodes: dict[str, Node],
    conduits: list[Conduit],
    diameter_scale: float,
    cross_section_segments: int,
):
    import plotly.graph_objects as go

    fig = go.Figure()

    visible_nodes = [node for node in nodes.values() if node.has_xyz]
    for kind, color, label in (("junction", "black", "Junctions"), ("outfall", "red", "Outfalls")):
        for fixed, symbol, suffix in ((False, "circle", "NotFixed spheres"), (True, "square", "Fixed cubes")):
            group = [
                node
                for node in visible_nodes
                if node.kind == kind and (node.node_type.strip().lower() == "fixed") == fixed
            ]
            fig.add_trace(
                go.Scatter3d(
                    x=[node.x for node in group],
                    y=[node.y for node in group],
                    z=[node.elevation for node in group],
                    mode="markers+text",
                    marker={"size": 5, "color": color, "symbol": symbol},
                    text=[node.name for node in group],
                    textposition="top center",
                    name=f"{label} ({suffix})",
                    hovertemplate=(
                        "<b>%{text}</b><br>"
                        "Type=" + ("Fixed" if fixed else "NotFixed") + "<br>"
                        "X=%{x:.3f}<br>Y=%{y:.3f}<br>Z=%{z:.3f}<extra></extra>"
                    ),
                )
            )

    skipped: list[str] = []
    for conduit in conduits:
        start_node = nodes.get(conduit.from_node)
        end_node = nodes.get(conduit.to_node)
        if start_node is None or end_node is None or not start_node.has_xyz or not end_node.has_xyz:
            skipped.append(conduit.name)
            continue

        cross_section = conduit_cross_section(conduit, diameter_scale, cross_section_segments)
        x, y, z, i_idx, j_idx, k_idx = tube_mesh(node_xyz(start_node), node_xyz(end_node), cross_section)
        if not x:
            skipped.append(conduit.name)
            continue

        fig.add_trace(
            go.Mesh3d(
                x=x,
                y=y,
                z=z,
                i=i_idx,
                j=j_idx,
                k=k_idx,
                name=conduit.name,
                color=conduit_color_hex(conduit, nodes),
                opacity=0.55,
                hovertemplate=(
                    f"<b>{conduit.name}</b><br>"
                    f"{conduit.from_node} -> {conduit.to_node}<br>"
                    f"Length={conduit.length:g}<br>"
                    f"Shape={conduit.shape}<br>"
                    f"Geom1={conduit.geom1:g}, Geom2={conduit.geom2:g}, Geom3={conduit.geom3:g}<br>"
                    f"Distance/Length={conduit_distance_ratio(conduit, nodes) or 0:g}"
                    "<extra></extra>"
                ),
                showscale=False,
            )
        )

    for label, color in (
        ("Conduits: 0.9 < Euclidian / Length < 1.1", "#919191"),
        ("Conduits: Euclidian / Length << 0.9", "#FF0000"),
        ("Conduits: Euclidian / Length >> 1.1", "#0000FF"),
    ):
        fig.add_trace(
            go.Scatter3d(
                x=[None],
                y=[None],
                z=[None],
                mode="lines",
                line={"color": color, "width": 8},
                name=label,
                hoverinfo="skip",
                showlegend=True,
            )
        )

    fig.update_layout(
        title="Vue 3D du reseau SWMM",
        scene={
            "xaxis_title": "X-Coord",
            "yaxis_title": "Y-Coord",
            "zaxis_title": "Elevation",
            "aspectmode": "data",
        },
        legend={"itemsizing": "constant"},
        margin={"l": 0, "r": 0, "t": 45, "b": 0},
    )

    return fig, skipped


def obj_coordinates(x: float, y: float, z: float, swap_yz: bool) -> tuple[float, float, float]:
    if swap_yz:
        return (x, z, y)
    return (x, y, z)


def safe_obj_name(name: str) -> str:
    return "".join(char if char.isalnum() or char in "._-" else "_" for char in name)


def write_material(handle, name: str, rgb: tuple[int, int, int]) -> None:
    r, g, b = (value / 255.0 for value in rgb)
    handle.write(f"newmtl {name}\n")
    handle.write(f"Kd {r:.4f} {g:.4f} {b:.4f}\n")
    handle.write(f"Ka {r * 0.15:.4f} {g * 0.15:.4f} {b * 0.15:.4f}\n\n")


def write_obj_materials(mtl_path: Path, nodes: dict[str, Node], conduits: list[Conduit]) -> None:
    with mtl_path.open("w", encoding="utf-8") as handle:
        handle.write("# Materials for SWMM OBJ export\n")
        for conduit in conduits:
            write_material(handle, f"conduit_{safe_obj_name(conduit.name)}", conduit_color_rgb(conduit_distance_ratio(conduit, nodes)))
        write_material(handle, "junction", (0, 0, 0))
        write_material(handle, "outfall", (255, 0, 0))


def sphere_mesh(
    center: tuple[float, float, float],
    radius: float,
    slices: int = 12,
    stacks: int = 6,
) -> tuple[list[float], list[float], list[float], list[int], list[int], list[int]]:
    x: list[float] = []
    y: list[float] = []
    z: list[float] = []

    for stack in range(stacks + 1):
        phi = math.pi * stack / stacks
        ring_radius = radius * math.sin(phi)
        ring_z = radius * math.cos(phi)
        for slice_idx in range(slices):
            theta = 2.0 * math.pi * slice_idx / slices
            x.append(center[0] + ring_radius * math.cos(theta))
            y.append(center[1] + ring_radius * math.sin(theta))
            z.append(center[2] + ring_z)

    i_idx: list[int] = []
    j_idx: list[int] = []
    k_idx: list[int] = []
    for stack in range(stacks):
        for slice_idx in range(slices):
            next_slice = (slice_idx + 1) % slices
            a = stack * slices + slice_idx
            b = stack * slices + next_slice
            c = (stack + 1) * slices + slice_idx
            d = (stack + 1) * slices + next_slice
            i_idx.extend([a, b])
            j_idx.extend([c, d])
            k_idx.extend([b, c])

    return x, y, z, i_idx, j_idx, k_idx


def cube_mesh(
    center: tuple[float, float, float],
    half_size: float,
) -> tuple[list[float], list[float], list[float], list[int], list[int], list[int]]:
    cx, cy, cz = center
    h = half_size
    vertices = [
        (cx - h, cy - h, cz - h),
        (cx + h, cy - h, cz - h),
        (cx + h, cy + h, cz - h),
        (cx - h, cy + h, cz - h),
        (cx - h, cy - h, cz + h),
        (cx + h, cy - h, cz + h),
        (cx + h, cy + h, cz + h),
        (cx - h, cy + h, cz + h),
    ]
    faces = [
        (0, 1, 2), (0, 2, 3),
        (4, 6, 5), (4, 7, 6),
        (0, 4, 5), (0, 5, 1),
        (1, 5, 6), (1, 6, 2),
        (2, 6, 7), (2, 7, 3),
        (3, 7, 4), (3, 4, 0),
    ]
    x = [vertex[0] for vertex in vertices]
    y = [vertex[1] for vertex in vertices]
    z = [vertex[2] for vertex in vertices]
    i_idx = [face[0] for face in faces]
    j_idx = [face[1] for face in faces]
    k_idx = [face[2] for face in faces]
    return x, y, z, i_idx, j_idx, k_idx


def write_conduits_obj(
    obj_path: Path,
    nodes: dict[str, Node],
    conduits: list[Conduit],
    diameter_scale: float,
    cross_section_segments: int,
    swap_yz: bool,
    export_nodes: bool,
    node_radius: float,
) -> list[str]:
    """Export the generated conduit and node meshes as a Wavefront OBJ file."""
    skipped: list[str] = []
    vertex_offset = 0
    mtl_path = obj_path.with_suffix(".mtl")
    write_obj_materials(mtl_path, nodes, conduits)

    with obj_path.open("w", encoding="utf-8") as handle:
        handle.write("# SWMM conduit network exported by visualize_swmm_3d.py\n")
        if swap_yz:
            handle.write("# Coordinates: OBJ X=SWMM X-Coord, OBJ Y=SWMM Elevation, OBJ Z=SWMM Y-Coord\n")
        else:
            handle.write("# Coordinates: OBJ X=SWMM X-Coord, OBJ Y=SWMM Y-Coord, OBJ Z=SWMM Elevation\n")
        handle.write(f"mtllib {mtl_path.name}\n")

        for conduit in conduits:
            start_node = nodes.get(conduit.from_node)
            end_node = nodes.get(conduit.to_node)
            if start_node is None or end_node is None or not start_node.has_xyz or not end_node.has_xyz:
                skipped.append(conduit.name)
                continue

            cross_section = conduit_cross_section(
                conduit,
                diameter_scale,
                max(cross_section_segments, 4),
            )
            x, y, z, i_idx, j_idx, k_idx = tube_mesh(node_xyz(start_node), node_xyz(end_node), cross_section)
            if not x:
                skipped.append(conduit.name)
                continue

            safe_name = safe_obj_name(conduit.name)
            handle.write(f"\no {safe_name}\n")
            handle.write(f"usemtl conduit_{safe_name}\n")
            handle.write(
                f"# {conduit.name}: {conduit.from_node} -> {conduit.to_node}; "
                f"shape={conduit.shape}; geom1={conduit.geom1:g}; "
                f"geom2={conduit.geom2:g}; geom3={conduit.geom3:g}; "
                f"distance_length_ratio={conduit_distance_ratio(conduit, nodes) or 0:g}\n"
            )

            for vx, vy, vz in zip(x, y, z):
                ox, oy, oz = obj_coordinates(vx, vy, vz, swap_yz)
                handle.write(f"v {ox:.6f} {oy:.6f} {oz:.6f}\n")

            for face_i, face_j, face_k in zip(i_idx, j_idx, k_idx):
                handle.write(
                    f"f {face_i + 1 + vertex_offset} "
                    f"{face_j + 1 + vertex_offset} "
                    f"{face_k + 1 + vertex_offset}\n"
                )

            vertex_offset += len(x)

        if export_nodes:
            for node in nodes.values():
                if not node.has_xyz:
                    continue
                safe_name = safe_obj_name(node.name)
                material = "outfall" if node.kind == "outfall" else "junction"
                is_fixed = node.node_type.strip().lower() == "fixed"
                if is_fixed:
                    x, y, z, i_idx, j_idx, k_idx = cube_mesh(node_xyz(node), node_radius)
                    shape = "cube"
                else:
                    x, y, z, i_idx, j_idx, k_idx = sphere_mesh(node_xyz(node), node_radius)
                    shape = "sphere"

                handle.write(f"\no node_{safe_name}\n")
                handle.write(f"usemtl {material}\n")
                handle.write(
                    f"# {node.name}: swmm_type={node.kind}; type={node.node_type}; "
                    f"shape={shape}; elevation={node.elevation:g}\n"
                )
                for vx, vy, vz in zip(x, y, z):
                    ox, oy, oz = obj_coordinates(vx, vy, vz, swap_yz)
                    handle.write(f"v {ox:.6f} {oy:.6f} {oz:.6f}\n")

                for face_i, face_j, face_k in zip(i_idx, j_idx, k_idx):
                    handle.write(
                        f"f {face_i + 1 + vertex_offset} "
                        f"{face_j + 1 + vertex_offset} "
                        f"{face_k + 1 + vertex_offset}\n"
                    )

                vertex_offset += len(x)

    return skipped


def print_summary(nodes: dict[str, Node], conduits: list[Conduit]) -> None:
    visible_nodes = sum(1 for node in nodes.values() if node.has_xyz)
    junctions = sum(1 for node in nodes.values() if node.kind == "junction")
    outfalls = sum(1 for node in nodes.values() if node.kind == "outfall")
    print(f"Nodes: {len(nodes)} ({junctions} junctions, {outfalls} outfalls), with XYZ: {visible_nodes}")
    print(f"Conduits: {len(conduits)}")

    shapes: dict[str, int] = {}
    for conduit in conduits:
        shapes[conduit.shape] = shapes.get(conduit.shape, 0) + 1
    for shape, count in sorted(shapes.items()):
        print(f"  {shape}: {count}")


def config_args() -> argparse.Namespace:
    return argparse.Namespace(
        inp_file=Path(INP_FILE),
        coordinates_xlsx=Path(COORDINATES_XLSX) if COORDINATES_XLSX else None,
        output=Path(OUTPUT_HTML) if OUTPUT_HTML else None,
        obj_output=Path(OUTPUT_OBJ) if OUTPUT_OBJ else None,
        no_obj=not EXPORT_OBJ,
        obj_swap_yz=OBJ_SWAP_YZ,
        obj_export_nodes=OBJ_EXPORT_NODES,
        obj_node_radius=OBJ_NODE_RADIUS,
        diameter_scale=DIAMETER_SCALE,
        cross_section_segments=CROSS_SECTION_SEGMENTS,
        summary_only=SUMMARY_ONLY,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a 3D view of an EPA SWMM .inp network.")
    parser.add_argument("inp_file", type=Path, help="Path to the SWMM .inp file.")
    parser.add_argument(
        "--coordinates-xlsx",
        type=Path,
        default=None,
        help="Excel file containing node name, X-Coord and Y-Coord in columns A, B and C.",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output HTML file. Defaults to <inp_file>_3d.html.",
    )
    parser.add_argument(
        "--obj-output",
        type=Path,
        default=None,
        help="Output OBJ file for conduit and node meshes. Defaults to <inp_file>_network.obj.",
    )
    parser.add_argument(
        "--no-obj",
        action="store_true",
        help="Do not export conduit and node meshes to OBJ.",
    )
    parser.add_argument(
        "--obj-no-swap-yz",
        action="store_true",
        help="Keep OBJ axes as X=X-Coord, Y=Y-Coord, Z=Elevation.",
    )
    parser.add_argument(
        "--obj-no-nodes",
        action="store_true",
        help="Do not export nodes as OBJ spheres.",
    )
    parser.add_argument(
        "--obj-node-radius",
        type=float,
        default=25.0,
        help="Radius of exported node spheres in model units.",
    )
    parser.add_argument(
        "--diameter-scale",
        type=float,
        default=80.0,
        help="Visual scale applied to conduit dimensions. Use 1 for real dimensions.",
    )
    parser.add_argument(
        "--cross-section-segments",
        type=int,
        default=16,
        help="Number of segments for circular/oval cross-sections.",
    )
    parser.add_argument(
        "--summary-only",
        action="store_true",
        help="Only parse the file and print a summary.",
    )
    return parser.parse_args()


def resolve_path(path: Path) -> Path:
    if path.is_absolute():
        return path.resolve()
    return (SCRIPT_DIR / path).resolve()


def main() -> int:
    args = parse_args()
    return run(args)


def run_from_config() -> int:
    """Run with the editable values from the Spyder configuration block."""
    return run(config_args())


def run(args: argparse.Namespace) -> int:
    inp_path = resolve_path(args.inp_file)
    if not inp_path.exists():
        print(f"File not found: {inp_path}")
        return 2

    coordinates_xlsx_path = resolve_path(args.coordinates_xlsx) if args.coordinates_xlsx else inp_path.with_name(COORDINATES_XLSX)
    if not coordinates_xlsx_path.exists():
        print(f"Coordinates Excel file not found: {coordinates_xlsx_path}")
        return 2

    try:
        nodes, conduits = parse_inp(inp_path, coordinates_xlsx_path)
    except RuntimeError as exc:
        print(exc)
        print("Install it in Anaconda with: conda install openpyxl")
        return 1

    print(f"Node XY coordinates read from: {coordinates_xlsx_path}")
    print_summary(nodes, conduits)
    if args.summary_only:
        return 0

    obj_skipped: list[str] = []
    if not args.no_obj:
        obj_path = resolve_path(args.obj_output) if args.obj_output else inp_path.with_name(f"{inp_path.stem}_network.obj")
        obj_skipped = write_conduits_obj(
            obj_path,
            nodes,
            conduits,
            diameter_scale=args.diameter_scale,
            cross_section_segments=args.cross_section_segments,
            swap_yz=getattr(args, "obj_swap_yz", not getattr(args, "obj_no_swap_yz", False)),
            export_nodes=getattr(args, "obj_export_nodes", not getattr(args, "obj_no_nodes", False)),
            node_radius=args.obj_node_radius,
        )
        print(f"OBJ written to: {obj_path}")

    try:
        fig, skipped = build_plotly_figure(
            nodes,
            conduits,
            diameter_scale=args.diameter_scale,
            cross_section_segments=max(args.cross_section_segments, 4),
        )
    except ImportError:
        print("Plotly is not installed. Install it with: pip install plotly")
        return 1

    output_path = resolve_path(args.output) if args.output else inp_path.with_name(f"{inp_path.stem}_3d.html")
    fig.write_html(output_path, include_plotlyjs=True)
    print(f"3D HTML written to: {output_path}")
    skipped_names = sorted(set(skipped + obj_skipped))
    if skipped_names:
        print(f"Skipped {len(skipped_names)} conduits without complete node coordinates: {', '.join(skipped_names)}")
    return 0


if __name__ == "__main__":
    # In Spyder, run this file directly after editing INP_FILE above.
    # From a terminal, command line arguments are still supported.
    import sys

    if len(sys.argv) > 1:
        raise SystemExit(main())
    raise SystemExit(run_from_config())
