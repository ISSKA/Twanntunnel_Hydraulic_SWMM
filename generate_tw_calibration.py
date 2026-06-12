from __future__ import annotations

import json
import math
from collections import OrderedDict, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import load_workbook


BASE_STATIONS_DIR = Path(
    r"O:\Projets en cours\SCIENCE\SP_Twann_tunnel\1_Data\0_MESURES_STATIONS"
)
OUTPUT_DIR = Path(
    r"O:\Projets en cours\SCIENCE\Sci.387_N05TWT_Appui_ISSKA_GG\1_PRODUCTION\SWMM\CALIBRATION"
)
CORRELATION_DIR = Path(
    r"O:\Projets en cours\SCIENCE\Sci.387_N05TWT_Appui_ISSKA_GG\1_PRODUCTION\SWMM\MEASURED_HYDRAULIC_CORRELATIONS"
)
FLOW_OUTPUT_FILE = OUTPUT_DIR / "TW_Calibration_Flow_m3s.txt"
LEVEL_OUTPUT_FILE = OUTPUT_DIR / "TW_Calibration_Level_depth.txt"
INP_FILE = Path(r"D:\Users\ISSKA\Documents\GitHub\Twanntunnel_Hydraulic_SWMM\SWMM_Twannbach.inp")
MEASURES_2016_DIR = Path(
    r"O:\Projets en cours\SCIENCE\SP_Twann_tunnel\Sci291_Investigations_Avant_projet\2016\5_Mesures"
)
SONDIERSTOLLEN_FILE = MEASURES_2016_DIR / "Sondes_Sondierstollen_2016-2017.xlsx"
NODE_ALIASES = {
    "SS1": "Sondierstollen",
}


def parse_datetime(value: object) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value

    if isinstance(value, (int, float)):
        # Excel serial date, including the 1900 leap-year compatibility offset.
        return datetime(1899, 12, 30) + timedelta(days=float(value))

    text = str(value).strip()
    for fmt in (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass

    return None


def parse_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace("'", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def round_to_nearest_hour(date_time: datetime) -> datetime:
    rounded = date_time + timedelta(minutes=30)
    return rounded.replace(minute=0, second=0, microsecond=0)


def add_point(
    series: OrderedDict[str, dict[datetime, list[float]]],
    station: str,
    date_time: datetime,
    value: float,
) -> None:
    if station not in series:
        series[station] = defaultdict(list)
    series[station][round_to_nearest_hour(date_time)].append(value)


def iter_sheet_values(
    workbook_path: Path,
    sheet_name: str | int,
    date_column: int,
    value_column: int,
) -> Iterable[tuple[datetime, float]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if isinstance(sheet_name, int):
            worksheet = workbook.worksheets[sheet_name]
        else:
            worksheet = workbook[sheet_name]

        max_column = max(date_column, value_column)
        for row in worksheet.iter_rows(min_row=2, max_col=max_column, values_only=True):
            date_time = parse_datetime(row[date_column - 1])
            value = parse_float(row[value_column - 1])
            if date_time is not None and value is not None:
                yield date_time, value
    finally:
        workbook.close()


def add_excel_series(
    series: OrderedDict[str, dict[datetime, list[float]]],
    workbook_path: Path,
    station: str,
    date_column: int,
    value_column: int,
    convert_value: Callable[[float], float],
    sheet_name: str | int = 0,
) -> None:
    for date_time, raw_value in iter_sheet_values(
        workbook_path, sheet_name, date_column, value_column
    ):
        add_point(series, station, date_time, convert_value(raw_value))


def read_swmm_node_elevations(inp_file: Path) -> dict[str, float]:
    elevations: dict[str, float] = {}
    section = ""

    for line in inp_file.read_text(encoding="cp1252").splitlines():
        stripped = line.strip()
        if stripped.startswith("[") and stripped.endswith("]"):
            section = stripped.strip("[]")
            continue
        if section not in {"JUNCTIONS", "OUTFALLS"}:
            continue
        if not stripped or stripped.startswith(";"):
            continue

        parts = stripped.split()
        if len(parts) >= 2:
            elevations[parts[0]] = float(parts[1])

    return elevations


def node_elevation(elevations: dict[str, float], station: str) -> float:
    node_name = NODE_ALIASES.get(station, station)
    try:
        return elevations[node_name]
    except KeyError as exc:
        raise KeyError(
            f"Altitude du noeud {node_name!r} introuvable dans {INP_FILE} "
            f"pour la station {station!r}"
        ) from exc


def hourly_average(
    series: OrderedDict[str, dict[datetime, list[float]]],
    station: str,
    hour: datetime,
) -> float | None:
    if station not in series or hour not in series[station]:
        return None

    values = series[station][hour]
    return sum(values) / len(values)


def valid_hours_without_large_gaps(
    series: OrderedDict[str, dict[datetime, list[float]]],
    station: str,
    max_gap: timedelta = timedelta(hours=48),
) -> set[datetime]:
    hours = sorted(series.get(station, {}))
    valid: set[datetime] = set()

    for index, hour in enumerate(hours):
        previous_ok = index == 0 or hour - hours[index - 1] <= max_gap
        next_ok = index == len(hours) - 1 or hours[index + 1] - hour <= max_gap
        if previous_ok and next_ok:
            valid.add(hour)

    return valid


def twannbach_oben_level_to_flow(level: float) -> float:
    flow_lps = (
        113759.0 * level**5
        - 182008.0 * level**4
        + 104604.0 * level**3
        - 20741.0 * level**2
        + 1430.1 * level
    )
    return flow_lps / 1000.0


def add_wasserhooliloch_flow(
    flow_series: OrderedDict[str, dict[datetime, list[float]]],
    derived_inputs: OrderedDict[str, dict[datetime, list[float]]],
) -> None:
    target_station = "Hooliloch_L"

    for hour in sorted(derived_inputs["Twannbach_Unten"]):
        q_unten = hourly_average(derived_inputs, "Twannbach_Unten", hour)
        q_oben = hourly_average(derived_inputs, "Twannbach_Oben", hour)
        q_fenster = hourly_average(derived_inputs, "Fensterstollen", hour)

        if q_unten is None or q_oben is None or q_fenster is None:
            continue

        q_wasserhooliloch = q_unten - (q_oben + q_fenster)
        if q_wasserhooliloch < 0:
            continue

        add_point(flow_series, target_station, hour, q_wasserhooliloch)


def add_sauser_flow(
    flow_series: OrderedDict[str, dict[datetime, list[float]]],
    derived_inputs: OrderedDict[str, dict[datetime, list[float]]],
) -> None:
    target_station = "Sauser_L"
    valid_aval = valid_hours_without_large_gaps(derived_inputs, "Twannbach_Unten")
    valid_amont = valid_hours_without_large_gaps(derived_inputs, "Twannbach_Oben")

    for hour in sorted(derived_inputs["Twannbach_Unten"]):
        q_aval = hourly_average(derived_inputs, "Twannbach_Unten", hour)
        q_amont = hourly_average(derived_inputs, "Twannbach_Oben", hour)
        q_fenster = hourly_average(derived_inputs, "Fensterstollen", hour)

        if q_aval is None or q_amont is None or q_fenster is None:
            continue
        if hour not in valid_aval or hour not in valid_amont:
            continue
        if q_amont > 0:
            continue

        q_sauser = q_aval - q_fenster
        if q_sauser > 0.5:
            continue

        add_point(flow_series, target_station, hour, q_sauser)


def averaged_series(
    series: OrderedDict[str, dict[datetime, list[float]]],
    station: str,
) -> dict[datetime, float]:
    return {
        hour: sum(values) / len(values)
        for hour, values in series.get(station, {}).items()
        if values
    }


def pearson_r(points: list[dict[str, object]]) -> float | None:
    if len(points) < 2:
        return None

    xs = [float(point["x"]) for point in points]
    ys = [float(point["y"]) for point in points]
    x_mean = sum(xs) / len(xs)
    y_mean = sum(ys) / len(ys)
    numerator = sum((x - x_mean) * (y - y_mean) for x, y in zip(xs, ys))
    x_var = sum((x - x_mean) ** 2 for x in xs)
    y_var = sum((y - y_mean) ** 2 for y in ys)
    if x_var == 0 or y_var == 0:
        return None

    return numerator / math.sqrt(x_var * y_var)


def write_sauser_correlation(
    flow_series: OrderedDict[str, dict[datetime, list[float]]],
) -> None:
    sauser = averaged_series(flow_series, "Sauser_L")
    brunn_source: OrderedDict[str, dict[datetime, list[float]]] = OrderedDict()
    add_excel_series(
        brunn_source,
        BASE_STATIONS_DIR
        / "Brunnmuehle_Quellteich"
        / "Brunnmuehle_Quellteich_hour_avrg.xlsx",
        "Brunnmuehle_Quellteich",
        date_column=1,
        value_column=2,
        convert_value=lambda value: value * 0.001,
    )
    brunn = averaged_series(brunn_source, "Brunnmuehle_Quellteich")
    points: list[dict[str, object]] = []

    for hour in sorted(set(sauser) & set(brunn)):
        points.append(
            {
                "date": hour.strftime("%Y-%m-%d %H:%M"),
                "year": hour.year,
                "x": round(brunn[hour], 6),
                "y": round(sauser[hour], 6),
            }
        )

    CORRELATION_DIR.mkdir(parents=True, exist_ok=True)
    stem = "Q_Sauser_L_vs_Q_Brunnmuehle_Quellteich"
    csv_path = CORRELATION_DIR / f"{stem}.csv"
    html_path = CORRELATION_DIR / f"{stem}.html"

    csv_lines = ["date;year;Q_Brunnmuehle_Quellteich_m3s;Q_Sauser_L_m3s"]
    for point in points:
        csv_lines.append(
            f"{point['date']};{point['year']};{point['x']};{point['y']}"
        )
    csv_path.write_text("\n".join(csv_lines) + "\n", encoding="cp1252")

    r_value = pearson_r(points)
    r_text = "NA" if r_value is None else f"{r_value:.3f}"
    years = sorted({int(point["year"]) for point in points})
    points_json = json.dumps(points, ensure_ascii=False)
    years_json = json.dumps(years)

    html = f"""<!doctype html>
<html lang="fr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{stem}</title>
<style>
  :root {{ color-scheme: light; font-family: Arial, Helvetica, sans-serif; color: #172026; background: #f4f6f7; }}
  body {{ margin: 0; }}
  main {{ max-width: 1220px; margin: 0 auto; padding: 24px; }}
  h1 {{ margin: 0 0 6px; font-size: 24px; line-height: 1.2; }}
  .meta {{ margin: 0 0 18px; color: #4f5d66; font-size: 14px; }}
  .layout {{ display: grid; grid-template-columns: 220px 1fr; gap: 18px; align-items: start; }}
  .panel {{ background: #fff; border: 1px solid #d9e0e4; border-radius: 8px; padding: 14px; }}
  .years {{ display: grid; gap: 7px; max-height: 620px; overflow: auto; }}
  .years label {{ display: flex; align-items: center; gap: 8px; font-size: 14px; }}
  .actions {{ display: flex; gap: 8px; margin-bottom: 12px; }}
  button {{ border: 1px solid #b8c3ca; border-radius: 6px; background: #fff; padding: 7px 10px; cursor: pointer; }}
  button:hover {{ background: #eef3f5; }}
  canvas {{ width: 100%; height: 680px; background: #fff; border: 1px solid #d9e0e4; border-radius: 8px; display: block; }}
  .hint {{ margin-top: 10px; color: #52616b; font-size: 13px; }}
  @media (max-width: 760px) {{
    main {{ padding: 14px; }}
    .layout {{ grid-template-columns: 1fr; }}
    canvas {{ height: 520px; }}
  }}
</style>
</head>
<body>
<main>
  <h1>{stem}</h1>
  <p class="meta">Points horaires communs. N = {len(points)}, r = {r_text}</p>
  <div class="layout">
    <aside class="panel">
      <div class="actions">
        <button type="button" id="all">Tout</button>
        <button type="button" id="none">Aucun</button>
      </div>
      <div class="years" id="years"></div>
      <p class="hint">X : Q_Brunnmuehle_Quellteich [m3/s]<br>Y : Q_Sauser_L [m3/s]</p>
    </aside>
    <canvas id="chart" width="980" height="680"></canvas>
  </div>
</main>
<script>
const points = {points_json};
const years = {years_json};
const active = new Set(years);
const colors = ["#1f77b4","#ff7f0e","#2ca02c","#d62728","#9467bd","#8c564b","#17becf","#bcbd22"];
const yearsBox = document.getElementById("years");
const canvas = document.getElementById("chart");
const ctx = canvas.getContext("2d");

function buildYears() {{
  yearsBox.innerHTML = "";
  years.forEach((year, i) => {{
    const label = document.createElement("label");
    const input = document.createElement("input");
    input.type = "checkbox";
    input.checked = true;
    input.dataset.year = year;
    input.addEventListener("change", () => {{
      if (input.checked) active.add(year); else active.delete(year);
      draw();
    }});
    const swatch = document.createElement("span");
    swatch.style.cssText = "display:inline-block;width:12px;height:12px;border-radius:50%;background:" + colors[i % colors.length];
    label.append(input, swatch, document.createTextNode(String(year)));
    yearsBox.appendChild(label);
  }});
}}

function scale(values, minPad = 0.06) {{
  let min = Math.min(...values);
  let max = Math.max(...values);
  if (min === max) {{ min -= 1; max += 1; }}
  const pad = (max - min) * minPad;
  return [Math.max(0, min - pad), max + pad];
}}

function draw() {{
  const selected = points.filter(p => active.has(p.year));
  ctx.clearRect(0, 0, canvas.width, canvas.height);
  ctx.fillStyle = "#ffffff";
  ctx.fillRect(0, 0, canvas.width, canvas.height);
  const margin = {{ left: 76, right: 28, top: 28, bottom: 64 }};
  const plotW = canvas.width - margin.left - margin.right;
  const plotH = canvas.height - margin.top - margin.bottom;
  if (!selected.length) return;
  const [xmin, xmax] = scale(selected.map(p => p.x));
  const [ymin, ymax] = scale(selected.map(p => p.y));
  const x = v => margin.left + ((v - xmin) / (xmax - xmin)) * plotW;
  const y = v => margin.top + plotH - ((v - ymin) / (ymax - ymin)) * plotH;

  ctx.strokeStyle = "#d9e0e4";
  ctx.lineWidth = 1;
  ctx.beginPath();
  ctx.moveTo(margin.left, margin.top);
  ctx.lineTo(margin.left, margin.top + plotH);
  ctx.lineTo(margin.left + plotW, margin.top + plotH);
  ctx.stroke();

  ctx.fillStyle = "#52616b";
  ctx.font = "13px Arial";
  ctx.textAlign = "center";
  ctx.fillText("Q_Brunnmuehle_Quellteich [m3/s]", margin.left + plotW / 2, canvas.height - 18);
  ctx.save();
  ctx.translate(18, margin.top + plotH / 2);
  ctx.rotate(-Math.PI / 2);
  ctx.fillText("Q_Sauser_L [m3/s]", 0, 0);
  ctx.restore();

  ctx.textAlign = "right";
  for (let i = 0; i <= 5; i++) {{
    const xv = xmin + (xmax - xmin) * i / 5;
    const yv = ymin + (ymax - ymin) * i / 5;
    ctx.fillText(xv.toFixed(3), x(xv), margin.top + plotH + 20);
    ctx.fillText(yv.toFixed(3), margin.left - 8, y(yv) + 4);
  }}

  selected.forEach(p => {{
    const idx = years.indexOf(p.year);
    ctx.fillStyle = colors[idx % colors.length];
    ctx.globalAlpha = 0.72;
    ctx.beginPath();
    ctx.arc(x(p.x), y(p.y), 3, 0, Math.PI * 2);
    ctx.fill();
  }});
  ctx.globalAlpha = 1;
}}

document.getElementById("all").addEventListener("click", () => {{
  active.clear(); years.forEach(y => active.add(y));
  document.querySelectorAll("#years input").forEach(i => i.checked = true);
  draw();
}});
document.getElementById("none").addEventListener("click", () => {{
  active.clear();
  document.querySelectorAll("#years input").forEach(i => i.checked = false);
  draw();
}});

buildYears();
draw();
</script>
</body>
</html>
"""
    html_path.write_text(html, encoding="utf-8")


def find_schuettstein_file() -> Path:
    matches = sorted(
        list(MEASURES_2016_DIR.glob("Sondes_Schutstein_Gischeren.xlsx"))
        + list(MEASURES_2016_DIR.glob("Sonde_Sch*tstein.xlsx"))
    )
    if not matches:
        raise FileNotFoundError(
            f"Fichier Sonde_Sch*tstein.xlsx introuvable dans {MEASURES_2016_DIR}"
        )
    return matches[0]


def format_number(value: float) -> str:
    return f"{value:.6f}".rstrip("0").rstrip(".")


def write_calibration_file(
    output_file: Path, series: OrderedDict[str, dict[datetime, list[float]]]
) -> None:
    lines = ["; Calibration files"]

    for station, hourly_values in series.items():
        lines.append(station)
        for hour in sorted(hourly_values):
            values = hourly_values[hour]
            average = sum(values) / len(values)
            lines.append(
                "\t".join(
                    (
                        "",
                        hour.strftime("%m/%d/%Y"),
                        hour.strftime("%H:%M:%S"),
                        format_number(average),
                    )
                )
            )

    output_file.parent.mkdir(parents=True, exist_ok=True)
    output_file.write_text("\n".join(lines) + "\n", encoding="cp1252")


def main() -> None:
    series_by_metric: OrderedDict[
        str, OrderedDict[str, dict[datetime, list[float]]]
    ] = OrderedDict(
        (
            ("Flow", OrderedDict()),
            ("Level", OrderedDict()),
        )
    )
    derived_inputs: OrderedDict[str, dict[datetime, list[float]]] = OrderedDict()
    node_elevations = read_swmm_node_elevations(INP_FILE)

    hourly_stations = (
        ("Brunnmuehle_Quelle", "Brunn_Teich_L", "Flow", 0.001),
        ("Entwaesserungstollen", "Entw_Sto_L", "Flow", 0.001),
        ("Fensterstollen", "Fenster_L", "Flow", 1.0),
        ("Wasserhooliloch_Sonde_2", "Holiloch_sonde", "Level", 1.0),
    )

    for source_station, target_station, metric, factor in hourly_stations:
        workbook_path = (
            BASE_STATIONS_DIR
            / source_station
            / f"{source_station}_hour_avrg.xlsx"
        )
        elevation = node_elevation(node_elevations, target_station) if metric == "Level" else 0.0
        add_excel_series(
            series_by_metric[metric],
            workbook_path,
            target_station,
            date_column=1,
            value_column=2,
            convert_value=lambda value, factor=factor, elevation=elevation: value
            * factor
            - elevation,
        )
        if source_station == "Fensterstollen":
            add_excel_series(
                derived_inputs,
                workbook_path,
                "Fensterstollen",
                date_column=1,
                value_column=2,
                convert_value=lambda value, factor=factor: value * factor,
            )

    add_excel_series(
        derived_inputs,
        BASE_STATIONS_DIR / "Twannbach_Unten" / "Twannbach_Unten_hour_avrg.xlsx",
        "Twannbach_Unten",
        date_column=1,
        value_column=2,
        convert_value=lambda value: value,
    )
    add_excel_series(
        derived_inputs,
        BASE_STATIONS_DIR / "Twannbach_Oben" / "Twannbach_Oben_hour_avrg.xlsx",
        "Twannbach_Oben",
        date_column=1,
        value_column=2,
        convert_value=twannbach_oben_level_to_flow,
    )
    add_wasserhooliloch_flow(series_by_metric["Flow"], derived_inputs)
    add_sauser_flow(series_by_metric["Flow"], derived_inputs)

    mbar_to_meter = 1.0 / 98.0665
    sondierstollen_sheets = {
        "SS1": (2, 3, 443.0),
        "SS4": (3, 5, 442.0),
        "SS5": (9, 11, 438.9),
        "SS6": (3, 5, 442.0),
    }

    for station, (date_column, value_column, altitude) in sondierstollen_sheets.items():
        elevation = node_elevation(node_elevations, station)
        add_excel_series(
            series_by_metric["Level"],
            SONDIERSTOLLEN_FILE,
            station,
            date_column=date_column,
            value_column=value_column,
            convert_value=lambda value, altitude=altitude, elevation=elevation: altitude
            + value * mbar_to_meter
            - elevation,
            sheet_name=station,
        )

    schuettstein_file = find_schuettstein_file()
    schuettstein_elevation = node_elevation(node_elevations, "SondeSchuettstein")
    add_excel_series(
        series_by_metric["Level"],
        schuettstein_file,
        "SondeSchuettstein",
        date_column=3,
        value_column=9,
        convert_value=lambda value: value - schuettstein_elevation,
    )
    gischeren_elevation = node_elevation(node_elevations, "SondeGischeren")
    add_excel_series(
        series_by_metric["Level"],
        schuettstein_file,
        "SondeGischeren",
        date_column=14,
        value_column=15,
        convert_value=lambda value: value - gischeren_elevation,
    )

    write_calibration_file(FLOW_OUTPUT_FILE, series_by_metric["Flow"])
    write_calibration_file(LEVEL_OUTPUT_FILE, series_by_metric["Level"])
    write_sauser_correlation(series_by_metric["Flow"])

    for metric, series in series_by_metric.items():
        for station, hourly_values in series.items():
            print(f"{metric} - {station}: {len(hourly_values)} valeurs horaires")
    print(f"Fichier debits ecrit: {FLOW_OUTPUT_FILE}")
    print(f"Fichier hauteurs ecrit: {LEVEL_OUTPUT_FILE}")


if __name__ == "__main__":
    main()
