from __future__ import annotations

import shutil
from dataclasses import dataclass
from datetime import datetime, timedelta
from html import escape
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


BASE_DIR = Path(r"O:\Projets en cours\SCIENCE\SP_Twann_tunnel\1_Data\0_MESURES_STATIONS")
OUTPUT_DIR = Path(r"O:\Projets en cours\SCIENCE\Sci.387_N05TWT_Appui_ISSKA_GG\1_PRODUCTION\SWMM\INPUT")
OUTPUT_FILE = OUTPUT_DIR / "Discharge_Input_SWMM.txt"
SYSTEME_EST_FILE = OUTPUT_DIR / "Discharge_Input_SWMM_Systeme_Est.txt"
PLOT_FILE = OUTPUT_DIR / "Discharge_Input_SWMM_plot.html"
FLOW_DURATION_CSV = OUTPUT_DIR / "Discharge_Input_SWMM_flow_duration_curve.csv"
FLOW_DURATION_HTML = OUTPUT_DIR / "Discharge_Input_SWMM_flow_duration_curve.html"
PLOTS_DIR = Path(r"D:\Users\ISSKA\Documents\GitHub\Twanntunnel_Hydraulic_SWMM\PLOTS")

MAX_ALLOWED_GAP_HOURS = 48
HOURS_PER_YEAR = 8760.0


@dataclass(frozen=True)
class Source:
    name: str
    path: Path
    converter: Callable[[float], float]
    sign: int = 1


def liters_per_second_to_m3_per_second(value: float) -> float:
    return value / 1000.0


def identity_m3_per_second(value: float) -> float:
    return value


def twannbach_oben_level_to_m3_per_second(level_m: float) -> float:
    x = level_m
    liters_per_second = (
        113759.0 * x**5
        - 182008.0 * x**4
        + 104604.0 * x**3
        - 20741.0 * x**2
        + 1430.1 * x
    )
    return liters_per_second / 1000.0


SOURCES = [
    Source(
        name="Brunnmuehle_Quelle",
        path=BASE_DIR / r"Brunnmuehle_Quelle\Brunnmuehle_Quelle_hour_avrg.xlsx",
        converter=liters_per_second_to_m3_per_second,
    ),
    Source(
        name="Entwaesserungstollen",
        path=BASE_DIR / r"Entwaesserungstollen\Entwaesserungstollen_hour_avrg.xlsx",
        converter=liters_per_second_to_m3_per_second,
    ),
    Source(
        name="Twannbach_Unten",
        path=BASE_DIR / r"Twannbach_Unten\Twannbach_Unten_hour_avrg.xlsx",
        converter=identity_m3_per_second,
    ),
    Source(
        name="Twannbach_Oben",
        path=BASE_DIR / r"Twannbach_Oben\Twannbach_Oben_hour_avrg.xlsx",
        converter=twannbach_oben_level_to_m3_per_second,
        sign=-1,
    ),
]


def normalize_hour(value: datetime) -> datetime:
    return datetime(value.year, value.month, value.day, value.hour)


def parse_excel_datetime(value) -> datetime:
    if isinstance(value, datetime):
        return normalize_hour(value)
    if isinstance(value, (int, float)):
        return normalize_hour(from_excel(value))
    raise ValueError(f"Date Excel non reconnue: {value!r}")


def read_series(source: Source) -> dict[datetime, float]:
    workbook = load_workbook(source.path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        series: dict[datetime, float] = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            raw_date, raw_value = row[:2]
            if raw_date is None or raw_value is None:
                continue

            date = parse_excel_datetime(raw_date)
            value = source.converter(float(raw_value))
            series[date] = value

        return series
    finally:
        workbook.close()


def iter_hours(start: datetime, end: datetime):
    current = start
    while current <= end:
        yield current
        current += timedelta(hours=1)


def fill_short_gaps_hourly(
    series: dict[datetime, float],
    start: datetime,
    end: datetime,
    max_allowed_gap_hours: int = MAX_ALLOWED_GAP_HOURS,
) -> dict[datetime, float]:
    hourly = dict(series)
    known_dates = sorted(date for date in series if start <= date <= end)

    for previous_date, next_date in zip(known_dates, known_dates[1:]):
        gap_hours = int((next_date - previous_date).total_seconds() / 3600) - 1
        if gap_hours <= 0 or gap_hours > max_allowed_gap_hours:
            continue

        previous_value = series[previous_date]
        next_value = series[next_date]

        for step in range(1, gap_hours + 1):
            current = previous_date + timedelta(hours=step)
            ratio = step / (gap_hours + 1)
            hourly[current] = previous_value + ratio * (next_value - previous_value)

    return hourly


def find_significant_gaps(
    series: dict[datetime, float],
    start: datetime,
    end: datetime,
    max_allowed_gap_hours: int = MAX_ALLOWED_GAP_HOURS,
) -> list[tuple[datetime, datetime, int]]:
    gaps: list[tuple[datetime, datetime, int]] = []
    missing_start: datetime | None = None

    for current in iter_hours(start, end):
        is_missing = current not in series

        if is_missing and missing_start is None:
            missing_start = current
        elif not is_missing and missing_start is not None:
            missing_end = current - timedelta(hours=1)
            hours = int((missing_end - missing_start).total_seconds() / 3600) + 1
            if hours > max_allowed_gap_hours:
                gaps.append((missing_start, missing_end, hours))
            missing_start = None

    if missing_start is not None:
        missing_end = end
        hours = int((missing_end - missing_start).total_seconds() / 3600) + 1
        if hours > max_allowed_gap_hours:
            gaps.append((missing_start, missing_end, hours))

    return gaps


def combine_flows(values: dict[str, float]) -> float:
    twannbach_unten = values["Twannbach_Unten"]
    twannbach_oben = values["Twannbach_Oben"]

    flow = (
        values["Brunnmuehle_Quelle"]
        + values["Entwaesserungstollen"]
        + twannbach_unten
    )

    if twannbach_oben <= twannbach_unten:
        flow -= twannbach_oben

    return flow


def format_number(value: float, decimals: int = 9) -> str:
    return f"{value:.{decimals}f}".rstrip("0").rstrip(".")


def copy_html_to_plots(output_file: Path) -> None:
    PLOTS_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(output_file, PLOTS_DIR / output_file.name)


def write_swmm_timeseries(
    sources: list[Source],
    all_series: dict[str, dict[datetime, float]],
    start: datetime,
    end: datetime,
    output_file: Path,
) -> tuple[int, int]:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    rows_written = 0
    twannbach_oben_ignored = 0

    with output_file.open("w", encoding="mbcs", newline="\r\n") as handle:
        for current in iter_hours(start, end):
            if not all(current in all_series[source.name] for source in sources):
                continue

            values = {
                source.name: all_series[source.name][current]
                for source in sources
            }
            if values["Twannbach_Oben"] > values["Twannbach_Unten"]:
                twannbach_oben_ignored += 1
            flow = combine_flows(values)
            handle.write(f"{current:%m/%d/%Y %H:%M}\t{flow:.9f}\n")
            rows_written += 1

    return rows_written, twannbach_oben_ignored


def write_scaled_timeseries(
    input_file: Path,
    output_file: Path,
    scale: float,
) -> int:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    rows_written = 0

    with input_file.open("r", encoding="mbcs") as source, output_file.open(
        "w",
        encoding="mbcs",
        newline="\r\n",
    ) as target:
        for line in source:
            line = line.strip()
            if not line:
                continue

            date_text, time_text, value_text = line.split(maxsplit=2)
            scaled_value = float(value_text) * scale
            target.write(f"{date_text} {time_text}\t{scaled_value:.9f}\n")
            rows_written += 1

    return rows_written


def read_generated_timeseries(path: Path) -> list[tuple[datetime, float]]:
    rows: list[tuple[datetime, float]] = []

    with path.open("r", encoding="mbcs") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue

            date_text, time_text, value_text = line.split(maxsplit=2)
            date = datetime.strptime(f"{date_text} {time_text}", "%m/%d/%Y %H:%M")
            rows.append((date, float(value_text)))

    return sorted(rows)


def detect_output_gaps(
    rows: list[tuple[datetime, float]],
) -> list[tuple[datetime, datetime, int]]:
    gaps: list[tuple[datetime, datetime, int]] = []

    for (previous_date, _), (current_date, _) in zip(rows, rows[1:]):
        missing_hours = int((current_date - previous_date).total_seconds() / 3600) - 1
        if missing_hours > 0:
            gaps.append(
                (
                    previous_date + timedelta(hours=1),
                    current_date - timedelta(hours=1),
                    missing_hours,
                )
            )

    return gaps


def build_svg_path(
    rows: list[tuple[datetime, float]],
    start: datetime,
    end: datetime,
    min_flow: float,
    max_flow: float,
    layout: dict[str, int],
) -> list[str]:
    left = layout["left"]
    top = layout["top"]
    plot_width = layout["plot_width"]
    plot_height = layout["plot_height"]
    total_seconds = max((end - start).total_seconds(), 1.0)
    flow_range = max(max_flow - min_flow, 0.000001)

    def x_position(date: datetime) -> float:
        return left + ((date - start).total_seconds() / total_seconds) * plot_width

    def y_position(flow: float) -> float:
        return top + ((max_flow - flow) / flow_range) * plot_height

    segments: list[str] = []
    current_segment: list[str] = []

    for index, (date, flow) in enumerate(rows):
        if index > 0 and (date - rows[index - 1][0]).total_seconds() > 3600 * 1.1:
            if len(current_segment) > 1:
                segments.append("M " + " L ".join(current_segment))
            current_segment = []

        current_segment.append(f"{x_position(date):.2f} {y_position(flow):.2f}")

    if len(current_segment) > 1:
        segments.append("M " + " L ".join(current_segment))

    return segments


def write_timeseries_plot(input_file: Path, output_file: Path) -> tuple[int, int]:
    rows = read_generated_timeseries(input_file)
    if not rows:
        raise RuntimeError(f"Aucune donnee lue dans {input_file}")

    gaps = detect_output_gaps(rows)
    start = rows[0][0]
    end = rows[-1][0]
    flows = [flow for _, flow in rows]
    min_flow = min(flows)
    max_flow = max(flows)

    width = 1600
    height = 900
    left = 95
    right = 35
    top = 60
    bottom = 115
    plot_width = width - left - right
    plot_height = height - top - bottom
    total_seconds = max((end - start).total_seconds(), 1.0)
    flow_range = max(max_flow - min_flow, 0.000001)
    layout = {
        "left": left,
        "top": top,
        "plot_width": plot_width,
        "plot_height": plot_height,
    }

    def x_position(date: datetime) -> float:
        return left + ((date - start).total_seconds() / total_seconds) * plot_width

    def y_position(flow: float) -> float:
        return top + ((max_flow - flow) / flow_range) * plot_height

    paths = build_svg_path(rows, start, end, min_flow, max_flow, layout)
    path_elements = "\n".join(
        f'<path d="{path}" fill="none" stroke="#1f6feb" '
        'stroke-width="1.4" stroke-linejoin="round" stroke-linecap="round" />'
        for path in paths
    )

    gap_rects = []
    for gap_start, gap_end, hours in gaps:
        x1 = x_position(gap_start)
        x2 = x_position(gap_end + timedelta(hours=1))
        gap_rects.append(
            f'<rect x="{x1:.2f}" y="{top}" width="{max(x2 - x1, 1):.2f}" '
            f'height="{plot_height}" fill="#d62828" opacity="0.18">'
            f"<title>Lacune: {gap_start:%Y-%m-%d %H:%M} - "
            f"{gap_end:%Y-%m-%d %H:%M} ({hours} h)</title></rect>"
        )

    x_ticks = []
    for index in range(9):
        date = start + timedelta(seconds=total_seconds * index / 8)
        x = x_position(date)
        x_ticks.append(
            f'<line x1="{x:.2f}" y1="{top}" x2="{x:.2f}" '
            f'y2="{top + plot_height}" stroke="#e5e7eb" />'
            f'<text x="{x:.2f}" y="{top + plot_height + 32}" '
            f'text-anchor="middle">{date:%Y-%m}</text>'
        )

    y_ticks = []
    for index in range(7):
        flow = min_flow + flow_range * index / 6
        y = y_position(flow)
        y_ticks.append(
            f'<line x1="{left}" y1="{y:.2f}" x2="{left + plot_width}" '
            f'y2="{y:.2f}" stroke="#e5e7eb" />'
            f'<text x="{left - 12}" y="{y + 4:.2f}" '
            f'text-anchor="end">{flow:.3f}</text>'
        )

    if gaps:
        gap_rows = "\n".join(
            f"<tr><td>{gap_start:%Y-%m-%d %H:%M}</td>"
            f"<td>{gap_end:%Y-%m-%d %H:%M}</td><td>{hours}</td></tr>"
            for gap_start, gap_end, hours in gaps
        )
    else:
        gap_rows = '<tr><td colspan="3">Aucune lacune detectee.</td></tr>'

    summary = (
        f"Fichier source: {escape(str(input_file))}<br>"
        f"Periode tracee: {start:%Y-%m-%d %H:%M} - {end:%Y-%m-%d %H:%M}<br>"
        f"Nombre de valeurs: {len(rows)}<br>"
        f"Nombre de lacunes: {len(gaps)}"
    )

    html = f"""<!doctype html>
<html lang="fr">
<head>
<meta charset="utf-8">
<title>Discharge Input SWMM - Time Series</title>
<style>
body {{ margin: 0; font-family: Arial, Helvetica, sans-serif; color: #172033; background: #f7f8fb; }}
main {{ max-width: 1720px; margin: 0 auto; padding: 28px 34px 44px; }}
h1 {{ margin: 0 0 8px; font-size: 24px; font-weight: 700; }}
.meta {{ margin-bottom: 18px; line-height: 1.5; color: #4b5563; font-size: 14px; }}
.figure {{ background: white; border: 1px solid #d9dee8; border-radius: 8px; padding: 18px; }}
svg {{ width: 100%; height: auto; display: block; }}
.axis text, text {{ font-size: 14px; fill: #374151; }}
.axis-title {{ font-size: 16px; font-weight: 700; fill: #172033; }}
.legend {{ display: flex; gap: 24px; align-items: center; margin: 14px 0 4px; color: #374151; font-size: 14px; }}
.swatch {{ width: 34px; height: 12px; display: inline-block; margin-right: 8px; vertical-align: -1px; }}
table {{ width: 100%; border-collapse: collapse; margin-top: 18px; background: white; border: 1px solid #d9dee8; }}
th, td {{ padding: 8px 10px; border-bottom: 1px solid #e5e7eb; text-align: left; font-size: 13px; }}
th {{ background: #eef2f7; }}
</style>
</head>
<body>
<main>
<h1>Discharge_Input_SWMM - Time series</h1>
<div class="meta">{summary}</div>
<div class="figure">
<svg viewBox="0 0 {width} {height}" role="img" aria-label="Time series SWMM avec lacunes">
<rect x="0" y="0" width="{width}" height="{height}" fill="#ffffff" />
<g class="axis">
{chr(10).join(x_ticks)}
{chr(10).join(y_ticks)}
<line x1="{left}" y1="{top}" x2="{left}" y2="{top + plot_height}" stroke="#172033" stroke-width="1.2" />
<line x1="{left}" y1="{top + plot_height}" x2="{left + plot_width}" y2="{top + plot_height}" stroke="#172033" stroke-width="1.2" />
<text class="axis-title" x="{left + plot_width / 2}" y="{height - 35}" text-anchor="middle">Date</text>
<text class="axis-title" x="24" y="{top + plot_height / 2}" transform="rotate(-90 24 {top + plot_height / 2})" text-anchor="middle">Debit [m3/s]</text>
</g>
<g>
{chr(10).join(gap_rects)}
</g>
<g>
{path_elements}
</g>
</svg>
<div class="legend">
<span><span class="swatch" style="background:#1f6feb"></span>Debit combine</span>
<span><span class="swatch" style="background:#d62828; opacity:.35"></span>Plage de lacune</span>
</div>
</div>
<table>
<thead><tr><th>Debut lacune</th><th>Fin lacune</th><th>Duree [h]</th></tr></thead>
<tbody>
{gap_rows}
</tbody>
</table>
</main>
</body>
</html>
"""
    output_file.write_text(html, encoding="utf-8")
    copy_html_to_plots(output_file)
    return len(rows), len(gaps)


def flow_duration_rows(
    rows: list[tuple[datetime, float]],
) -> list[dict[str, float | int]]:
    sorted_flows = sorted((flow for _, flow in rows), reverse=True)
    count = len(sorted_flows)
    duration_rows: list[dict[str, float | int]] = []

    for index, flow in enumerate(sorted_flows, start=1):
        exceedance_probability = index / (count + 1)
        duration_rows.append(
            {
                "Rank": index,
                "ExceedanceProbability": exceedance_probability,
                "ExceedancePercent": exceedance_probability * 100.0,
                "HoursPerYear": exceedance_probability * HOURS_PER_YEAR,
                "FlowM3s": flow,
            }
        )

    return duration_rows


def write_flow_duration_csv(
    output_file: Path,
    duration_rows: list[dict[str, float | int]],
) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    lines = ["Rank;ExceedanceProbability;ExceedancePercent;HoursPerYear;FlowM3s"]

    for row in duration_rows:
        lines.append(
            ";".join(
                (
                    str(row["Rank"]),
                    format_number(float(row["ExceedanceProbability"])),
                    format_number(float(row["ExceedancePercent"])),
                    format_number(float(row["HoursPerYear"])),
                    format_number(float(row["FlowM3s"])),
                )
            )
        )

    output_file.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")


def interpolated_flow_at_hours(
    duration_rows: list[dict[str, float | int]],
    hours_per_year: float,
) -> float:
    if hours_per_year <= float(duration_rows[0]["HoursPerYear"]):
        return float(duration_rows[0]["FlowM3s"])
    if hours_per_year >= float(duration_rows[-1]["HoursPerYear"]):
        return float(duration_rows[-1]["FlowM3s"])

    for previous, current in zip(duration_rows, duration_rows[1:]):
        previous_hours = float(previous["HoursPerYear"])
        current_hours = float(current["HoursPerYear"])
        if previous_hours <= hours_per_year <= current_hours:
            ratio = (hours_per_year - previous_hours) / (current_hours - previous_hours)
            previous_flow = float(previous["FlowM3s"])
            current_flow = float(current["FlowM3s"])
            return previous_flow + ratio * (current_flow - previous_flow)

    return float(duration_rows[-1]["FlowM3s"])


def downsample_duration_rows(
    duration_rows: list[dict[str, float | int]],
    max_points: int = 1800,
) -> list[dict[str, float | int]]:
    if len(duration_rows) <= max_points:
        return duration_rows

    step = max(1, len(duration_rows) // max_points)
    sampled = duration_rows[::step]
    if sampled[-1] is not duration_rows[-1]:
        sampled.append(duration_rows[-1])
    return sampled


def write_flow_duration_html(
    input_file: Path,
    output_file: Path,
    rows: list[tuple[datetime, float]],
    duration_rows: list[dict[str, float | int]],
) -> None:
    start = rows[0][0]
    end = rows[-1][0]
    max_hours = HOURS_PER_YEAR
    max_flow = max(float(row["FlowM3s"]) for row in duration_rows) * 1.05
    min_flow = 0.0
    flow_range = max(max_flow - min_flow, 0.000001)

    width = 1500
    height = 880
    left = 95
    right = 40
    top = 60
    bottom = 110
    plot_width = width - left - right
    plot_height = height - top - bottom

    def x_position(hours_per_year: float) -> float:
        return left + (hours_per_year / max_hours) * plot_width

    def y_position(flow: float) -> float:
        return top + ((max_flow - flow) / flow_range) * plot_height

    x_ticks = "\n".join(
        f'<line x1="{x_position(hours):.2f}" y1="{top}" x2="{x_position(hours):.2f}" '
        f'y2="{top + plot_height}" stroke="#e5e7eb" />'
        f'<text x="{x_position(hours):.2f}" y="{top + plot_height + 30}" text-anchor="middle">{hours:g}</text>'
        for hours in [0, 1000, 2000, 3000, 4000, 5000, 6000, 7000, 8000, 8760]
    )
    y_ticks = "\n".join(
        f'<line x1="{left}" y1="{y_position(flow):.2f}" x2="{left + plot_width}" '
        f'y2="{y_position(flow):.2f}" stroke="#e5e7eb" />'
        f'<text x="{left - 12}" y="{y_position(flow) + 4:.2f}" text-anchor="end">{format_number(flow, 3)}</text>'
        for flow in [min_flow + flow_range * index / 7 for index in range(8)]
    )
    path_points = [
        f'{x_position(float(row["HoursPerYear"])):.2f} {y_position(float(row["FlowM3s"])):.2f}'
        for row in downsample_duration_rows(duration_rows)
    ]
    duration_path = "M " + " L ".join(path_points)

    reference_hours = [1, 10, 100, 500, 1000, 2000, 4380, 7000, 8000]
    reference_rows = "\n".join(
        "<tr>"
        f"<td>{hours:g}</td>"
        f"<td>{format_number(hours / HOURS_PER_YEAR * 100.0, 3)}</td>"
        f"<td>{format_number(interpolated_flow_at_hours(duration_rows, hours), 3)}</td>"
        "</tr>"
        for hours in reference_hours
    )

    html = f"""<!doctype html>
<html lang="fr">
<head>
<meta charset="utf-8">
<title>Courbe des debits classes - Discharge Input SWMM</title>
<style>
body {{ margin:0; font-family:Arial, Helvetica, sans-serif; color:#172033; background:#f7f8fb; }}
main {{ max-width:1620px; margin:0 auto; padding:28px 34px 44px; }}
h1 {{ margin:0 0 8px; font-size:24px; }}
h2 {{ margin:26px 0 10px; font-size:18px; }}
.meta {{ margin-bottom:18px; line-height:1.55; color:#4b5563; font-size:14px; }}
.figure {{ background:white; border:1px solid #d9dee8; border-radius:8px; padding:18px; }}
svg {{ width:100%; height:auto; display:block; }}
text {{ font-size:14px; fill:#374151; }}
.axis-title {{ font-size:16px; font-weight:700; fill:#172033; }}
table {{ width:100%; border-collapse:collapse; margin-top:12px; background:white; border:1px solid #d9dee8; }}
th,td {{ padding:8px 10px; border-bottom:1px solid #e5e7eb; text-align:left; font-size:13px; }}
th {{ background:#eef2f7; }}
</style>
</head>
<body>
<main>
<h1>Courbe des debits classes - Discharge_Input_SWMM</h1>
<div class="meta">Fichier source: {escape(str(input_file))}<br>Pas de temps: horaire<br>Echelle: duree annuelle de depassement [h/an]<br>Periode: {start:%Y-%m-%d %H:%M} - {end:%Y-%m-%d %H:%M}<br>Nombre de valeurs horaires: {len(rows)}</div>
<div class="figure">
<svg viewBox="0 0 {width} {height}" role="img" aria-label="Courbe des debits classes">
<rect x="0" y="0" width="{width}" height="{height}" fill="#ffffff" />
<g>
{x_ticks}
{y_ticks}
<line x1="{left}" y1="{top}" x2="{left}" y2="{top + plot_height}" stroke="#172033" stroke-width="1.2" />
<line x1="{left}" y1="{top + plot_height}" x2="{left + plot_width}" y2="{top + plot_height}" stroke="#172033" stroke-width="1.2" />
<text class="axis-title" x="{left + plot_width / 2}" y="{height - 35}" text-anchor="middle">Duree annuelle de depassement [h/an]</text>
<text class="axis-title" x="24" y="{top + plot_height / 2}" transform="rotate(-90 24 {top + plot_height / 2})" text-anchor="middle">Debit [m3/s]</text>
</g>
<path d="{duration_path}" fill="none" stroke="#1f6feb" stroke-width="2.2" stroke-linejoin="round" stroke-linecap="round" />
</svg>
</div>
<h2>Valeurs reperes</h2>
<table>
<thead><tr><th>Duree depassee [h/an]</th><th>Frequence annuelle [%]</th><th>Debit classe [m3/s]</th></tr></thead>
<tbody>
{reference_rows}
</tbody>
</table>
</main>
</body>
</html>
"""
    output_file.write_text(html, encoding="utf-8")
    copy_html_to_plots(output_file)


def write_flow_duration_curve(
    input_file: Path,
    csv_file: Path,
    html_file: Path,
) -> int:
    rows = read_generated_timeseries(input_file)
    if not rows:
        raise RuntimeError(f"Aucune donnee lue dans {input_file}")

    duration_rows = flow_duration_rows(rows)
    write_flow_duration_csv(csv_file, duration_rows)
    write_flow_duration_html(input_file, html_file, rows, duration_rows)
    return len(duration_rows)


def main() -> None:
    raw_series: dict[str, dict[datetime, float]] = {}

    for source in SOURCES:
        print(f"Lecture {source.name}: {source.path}")
        series = read_series(source)
        if not series:
            raise RuntimeError(f"Aucune donnee lue pour {source.name}")

        raw_series[source.name] = series
        print(
            f"  {len(series)} valeurs, "
            f"{min(series):%Y-%m-%d %H:%M} -> {max(series):%Y-%m-%d %H:%M}"
        )

    start = max(min(series) for series in raw_series.values())
    end = min(max(series) for series in raw_series.values())

    hourly_series = {
        source.name: fill_short_gaps_hourly(raw_series[source.name], start, end)
        for source in SOURCES
    }

    rows_written, twannbach_oben_ignored = write_swmm_timeseries(
        SOURCES,
        hourly_series,
        start,
        end,
        OUTPUT_FILE,
    )
    systeme_est_rows = write_scaled_timeseries(
        OUTPUT_FILE,
        SYSTEME_EST_FILE,
        scale=0.1,
    )
    plot_rows, plot_gaps = write_timeseries_plot(OUTPUT_FILE, PLOT_FILE)
    flow_duration_count = write_flow_duration_curve(
        OUTPUT_FILE,
        FLOW_DURATION_CSV,
        FLOW_DURATION_HTML,
    )

    print()
    print(f"Fichier SWMM: {OUTPUT_FILE}")
    print(f"Fichier SWMM Systeme Est: {SYSTEME_EST_FILE}")
    print(f"Graphique HTML: {PLOT_FILE}")
    print(f"Courbe des debits classes HTML: {FLOW_DURATION_HTML}")
    print(f"Courbe des debits classes CSV: {FLOW_DURATION_CSV}")
    print(f"Periode commune: {start:%Y-%m-%d %H:%M} -> {end:%Y-%m-%d %H:%M}")
    print(f"Lignes ecrites: {rows_written}")
    print(f"Lignes Systeme Est ecrites: {systeme_est_rows}")
    print(f"Valeurs tracees: {plot_rows}")
    print(f"Lacunes tracees: {plot_gaps}")
    print(f"Valeurs courbe des debits classes: {flow_duration_count}")
    print("Pas de temps: horaire")
    print(f"Twannbach_Oben non soustrait: {twannbach_oben_ignored} heures")
    print()
    print(f"Lacunes > {MAX_ALLOWED_GAP_HOURS} h:")

    for source in SOURCES:
        gaps = find_significant_gaps(raw_series[source.name], start, end)
        print(f"- {source.name}: {len(gaps)}")
        for gap_start, gap_end, hours in gaps:
            print(f"  {gap_start:%Y-%m-%d %H:%M} -> {gap_end:%Y-%m-%d %H:%M} ({hours} h)")


if __name__ == "__main__":
    main()
